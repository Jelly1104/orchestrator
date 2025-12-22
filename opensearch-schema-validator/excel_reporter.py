"""
Excel Report Generator for Medigate Validation
"""
from datetime import datetime
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MATCH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_report(validation_results: List[Dict], output_path: str):
    """
    Create Excel report from validation results

    validation_results format:
    [
        {
            "content_id": 1192725,
            "url": "https://...",
            "content_type": "recruit",
            "status": "success" | "error",
            "error_message": "",
            "match_rate": 85.5,
            "total_fields": 20,
            "matched_fields": 17,
            "mismatched_fields": 3,
            "comparisons": [
                {"field": "...", "description": "...", "db_value": "...", "json_value": "...", "match": True/False}
            ]
        }
    ]
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "요약"
    create_summary_sheet(ws_summary, validation_results)

    # Sheet 2: Detail
    ws_detail = wb.create_sheet("상세비교")
    create_detail_sheet(ws_detail, validation_results)

    # Sheet 3: Mismatches Only
    ws_mismatch = wb.create_sheet("불일치항목")
    create_mismatch_sheet(ws_mismatch, validation_results)

    wb.save(output_path)
    print(f"Report saved: {output_path}")


def create_summary_sheet(ws, results: List[Dict]):
    """Create summary sheet"""
    headers = ["No", "Content ID", "URL", "타입", "상태", "일치율(%)", "전체필드", "일치", "불일치"]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Write data
    for row_idx, result in enumerate(results, 2):
        data = [
            row_idx - 1,
            result.get('content_id', 'N/A'),
            result.get('url', ''),
            result.get('content_type', ''),
            result.get('status', ''),
            result.get('match_rate', 0),
            result.get('total_fields', 0),
            result.get('matched_fields', 0),
            result.get('mismatched_fields', 0),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = THIN_BORDER

            # Color coding for status
            if col == 5:  # Status column
                if value == "success":
                    cell.fill = MATCH_FILL
                elif value == "error":
                    cell.fill = ERROR_FILL

            # Color coding for match rate
            if col == 6:  # Match rate
                if isinstance(value, (int, float)):
                    if value >= 90:
                        cell.fill = MATCH_FILL
                    elif value < 70:
                        cell.fill = MISMATCH_FILL

    # Adjust column widths
    column_widths = [5, 12, 50, 10, 10, 12, 10, 8, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_detail_sheet(ws, results: List[Dict]):
    """Create detail comparison sheet"""
    # Check if UI data is available
    ui_enabled = any(r.get('ui_enabled') for r in results)

    if ui_enabled:
        headers = ["Content ID", "필드(JSON)", "설명", "DB값", "JSON값", "UI값", "DB=JSON", "DB=UI", "DB컬럼"]
        column_widths = [12, 35, 15, 35, 35, 35, 8, 8, 25]
    else:
        headers = ["Content ID", "필드(JSON)", "설명", "DB값", "JSON값", "결과", "DB컬럼"]
        column_widths = [12, 35, 15, 40, 40, 8, 25]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    row_idx = 2
    for result in results:
        content_id = result.get('content_id', 'N/A')
        comparisons = result.get('comparisons', [])

        for comp in comparisons:
            if ui_enabled:
                data = [
                    content_id,
                    comp.get('field', ''),
                    comp.get('description', ''),
                    comp.get('db_value', ''),
                    comp.get('json_value', ''),
                    comp.get('ui_value', 'N/A'),
                    "O" if comp.get('match') else "X",
                    "O" if comp.get('ui_match_db') else ("X" if comp.get('ui_match_db') is False else "-"),
                    comp.get('db_column', ''),
                ]
            else:
                data = [
                    content_id,
                    comp.get('field', ''),
                    comp.get('description', ''),
                    comp.get('db_value', ''),
                    comp.get('json_value', ''),
                    "O" if comp.get('match') else "X",
                    comp.get('db_column', ''),
                ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = THIN_BORDER

                # Truncate long values
                if isinstance(value, str) and len(value) > 100:
                    cell.value = value[:100] + "..."

            # Color coding for match result
            if ui_enabled:
                # DB=JSON column
                match_cell = ws.cell(row=row_idx, column=7)
                if comp.get('match'):
                    match_cell.fill = MATCH_FILL
                else:
                    match_cell.fill = MISMATCH_FILL
                # DB=UI column
                ui_match_cell = ws.cell(row=row_idx, column=8)
                if comp.get('ui_match_db') is True:
                    ui_match_cell.fill = MATCH_FILL
                elif comp.get('ui_match_db') is False:
                    ui_match_cell.fill = MISMATCH_FILL
            else:
                match_cell = ws.cell(row=row_idx, column=6)
                if comp.get('match'):
                    match_cell.fill = MATCH_FILL
                else:
                    match_cell.fill = MISMATCH_FILL

            row_idx += 1

    # Adjust column widths
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def create_mismatch_sheet(ws, results: List[Dict]):
    """Create sheet with only mismatched fields"""
    # Check if UI data is available
    ui_enabled = any(r.get('ui_enabled') for r in results)

    if ui_enabled:
        headers = ["Content ID", "URL", "필드(JSON)", "설명", "DB값", "JSON값", "UI값", "불일치유형", "DB컬럼"]
        column_widths = [12, 40, 30, 15, 30, 30, 30, 15, 20]
    else:
        headers = ["Content ID", "URL", "필드(JSON)", "설명", "DB값", "JSON값", "DB컬럼"]
        column_widths = [12, 45, 35, 15, 40, 40, 25]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    row_idx = 2
    for result in results:
        content_id = result.get('content_id', 'N/A')
        url = result.get('url', '')
        comparisons = result.get('comparisons', [])

        for comp in comparisons:
            # Check for any mismatch: DB!=JSON or DB!=UI
            db_json_mismatch = not comp.get('match')
            db_ui_mismatch = comp.get('ui_match_db') is False

            if db_json_mismatch or db_ui_mismatch:
                # Determine mismatch type
                if db_json_mismatch and db_ui_mismatch:
                    mismatch_type = "DB!=JSON,UI"
                elif db_json_mismatch:
                    mismatch_type = "DB!=JSON"
                else:
                    mismatch_type = "DB!=UI"

                if ui_enabled:
                    data = [
                        content_id,
                        url,
                        comp.get('field', ''),
                        comp.get('description', ''),
                        comp.get('db_value', ''),
                        comp.get('json_value', ''),
                        comp.get('ui_value', 'N/A'),
                        mismatch_type,
                        comp.get('db_column', ''),
                    ]
                else:
                    data = [
                        content_id,
                        url,
                        comp.get('field', ''),
                        comp.get('description', ''),
                        comp.get('db_value', ''),
                        comp.get('json_value', ''),
                        comp.get('db_column', ''),
                    ]

                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = THIN_BORDER
                    cell.fill = MISMATCH_FILL

                    # Truncate long values
                    if isinstance(value, str) and len(value) > 100:
                        cell.value = value[:100] + "..."

                row_idx += 1

    # Adjust column widths
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Add message if no mismatches
    if row_idx == 2:
        msg = "불일치 항목이 없습니다."
        ws.cell(row=2, column=1, value=msg)
        merge_cols = len(headers)
        ws.merge_cells(f'A2:{get_column_letter(merge_cols)}2')


def generate_report_filename() -> str:
    """Generate report filename with timestamp in reports folder"""
    import os
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    reports_dir = os.path.join(script_dir, "reports")
    os.makedirs(reports_dir, exist_ok=True)
    return os.path.join(reports_dir, f"validation_report_{timestamp}.xlsx")
